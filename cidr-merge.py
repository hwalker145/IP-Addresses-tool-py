from netaddr import IPNetwork, cidr_merge
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook, Workbook
from pathlib import Path
import sys

# main code at the bottom

# this list will contain all runtime warnings as triggered by the functions
runtime_warnings = []

# this assigns Path objects into a list, so that it contains all
# paths which lead to .xlsx files inside the cwd
def getPaths(paths: list):
    # WIP dircount = 0

    # GOAL: find .xlsx files and add them to an array as a Path object
    # iterates over each file/directory in the cwd
    for child in Path.cwd().iterdir():
        # finds directory by the name 'xlsx'
        if(child.is_dir() and child.name == 'xlsx'):
            # WIP dircount += 1

            # iterates over each file/directory in the xlsx directory
            for grandchild in Path('xlsx').iterdir():
                # finds xlsx files, and appends them to a list
                if(grandchild.suffix == '.xlsx' and grandchild.name != 'Output.xlsx'):
                    paths.append(grandchild)
            # error check for no xlsx files in the xlsx directory
            # this is an error not a warning
            if(len(paths) == 0):
                print("No file in the directory 'xlsx' was in .xlsx form. Exiting...\n")
                exit

    # WIP warningHandler('00', dircount)

# reads all valid IP ranges into a list for v4 and one for v6
def readSheet(sheet: Worksheet, ranges_v4: list, ranges_v6: list):
    # the subnetcol variable refers to which column in the spread
    # sheet contains the subnets
    # searches through the columns to find the correct header
    subnetCol = findSubnetHeader(sheet)
    if(subnetCol == 0):
        return
    
    # reads through the 'Subnet' column
    for column in sheet.iter_cols(min_col=subnetCol, max_col=subnetCol):
        for cell in column:
            # converts to and verifies IP network
            net = stringToRange(cell.value)
            # appends to the list by version
            if(net):
                if(net.version == 4):
                    ranges_v4.append(net)
                elif(net.version == 6):
                    ranges_v6.append(net)

# function outputs a workbook with the aggregated addresses
def writeBook(workbook: Workbook, ranges_v4: list, ranges_v6: list, out_file_name: str):
    # creates a sheet
    sheet = workbook.create_sheet('Output')

    # labels the column headers
    sheet.cell(1,1).value = 'v4'
    sheet.cell(1,2).value = 'v6'

    # writes the ranges into the cells
    for row in range(len(ranges_v4)):
        sheet.cell(row+2, 1).value = str(ranges_v4[row])
    for row in range(len(ranges_v6)):
        sheet.cell(row+2, 2).value = str(ranges_v6[row])

    # saves as a workbook in xlsx directory
    workbook.save(out_file_name)

# searches through a worksheet object for a column header named 'Subnet'
def findSubnetHeader(sheet: Worksheet) -> int:
    # searches through all column headers to find 'Subnet'
    for cell in sheet[1]:
        if cell.value == 'Subnet':
            return cell.column
        if not cell.value:
            return 0

    # zero if no column header reads 'Subnet'
    return 0

# this function checks if the IP ranges are valid before constructing as IPNetwork()
def stringToRange(input: str) -> IPNetwork:
    # if the string contains neither . or :, then it's not an IP Range
    if(input.find('.') == -1 and input.find(':') == -1):
        return 0
    else:
        return IPNetwork(input)

# this part is WIP
"""
def warningHandler(context, data):
    if(context == '00'):
        if(data > 1):
            runtime_warnings.append({'context': '00', 
                                     'data': data,
                                     'less': 'Multiple directories named \'xlsx\'...', 
                                     'more': 'There were ' + data + ' directories by that name.'})
    if(context == '01'):
        if(data):
            runtime_warnings.append({'context': '01',
                                     'data': data,
                                     'less': 'There were sheet(s) which had no \'Subnet\' header.',
                                     'more': data + ' worksheets like this...'}) """

# globals
xlsx_paths = []
V4_list = []
V6_list = []

# main executed code here

# all applicable paths read into list
# getPaths(xlsx_paths)

if len(sys.argv) < 3:
    print("USAGE: \ninput.xlsx output.xlsx")

# since we are only working with one path
xlsx_paths.append(sys.argv[1])

# all cidr ranges read into lists
for filepath in xlsx_paths:
    wb = load_workbook(filepath)
    for sheetname in wb.sheetnames:
        readSheet(wb[sheetname], V4_list, V6_list)

out_name: str = 'Output.xlsx'

if len(sys.argv) >= 3:
    out_name = sys.argv[2]

# written into new workbook
wb = Workbook()
writeBook(wb, cidr_merge(V4_list), cidr_merge(V6_list), out_name)

# fin