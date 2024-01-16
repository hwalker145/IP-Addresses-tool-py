from netaddr import IPNetwork, all_matching_cidrs
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
import pandas as pd
import sys

# accepted header names
headers = [
    'subnets',
    'subnet',
    'cidrs',
    'cidr'
]

# checks if there is the correct amount of arguments on the CLI
if len(sys.argv) < 3:
    print("USAGE: \n<file> <address>")
    exit()

# checks the xlsx file in the CLI if it is really a .xlsx
if Path(sys.argv[1]).suffix != ".xlsx":
    print("USAGE: \n<file> <address>")
    exit()

# this function checks if the IP ranges are valid before constructing as IPNetwork()
def stringToRange(input: str) -> IPNetwork:
    # if the string contains neither . or :, then it's not an IP Range
    if(input.find('.') == -1 and input.find(':') == -1):
        return 0
    else:
        return IPNetwork(input)

def findHeader(sheet: Worksheet) -> int:
    for cell in sheet[1]:
        if cell.value in headers:
            return cell.column
        if not cell.value:
            return 0

def readSheet(sheet: Worksheet, ranges: list[IPNetwork]):
    # finds the column with the relevant data
    subnet_column: int = findHeader(sheet)
    # iterates through that column and adds IPNetwork objects into the list
    for column in sheet.iter_cols(subnet_column, subnet_column):
        for cell in column[1:]:
            # only adds to the list if it returns true (is not zero)
            if cell.value:
                # I call the stringtorange function to make sure each address
                # is a proper cidr address/range
                ranges.append({'index': cell.row, 'cidr': stringToRange(cell.value)})
    # returns final product
    return ranges

# this function compiles the matches and turns it into a df
def reportMatches(sheet: Worksheet, matches: list[dict]) -> None:
    # grabs the header values as values
    data = [[cell.value for cell in ws[1]]]

    # appends all the data into a list
    for m in matches:
        data.append([cell.value for cell in ws[m['index']]])

    # turns it into a df
    df = pd.DataFrame(data[1:], columns = data[0])

    # TODO is to figure out how much of the df to print
    # for now it's the whole thing
    print(df)

address = sys.argv[2]

wb: Workbook = load_workbook(sys.argv[1])

ranges: list[dict] = []

# calling the readSheet subroutine
for ws in wb.worksheets:
    readSheet(ws, ranges)

# calls the all_matching_cidrs function from netaddr
matches: list[IPNetwork] = []

# iterates through ranges to make matches
for range in ranges:
    if address in range['cidr']:
        matches.append(range)

# reports
print(str(len(matches)) + " matching cidr\nfor the given address.\n")
reportMatches(ws, matches)